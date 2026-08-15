# -*- coding: utf-8 -*-
"""点検メモ（`docs/点検メモ.md`）を作る —— パソコン上で書き込む表

**チートシートとは役割が違う。** チートシートは作図中に番号を引くためのもので、
よく使うものしか載せない。こちらは**`elements/` のぜんぶ**を並べて、
使っていて気づいたことを書き留めるためのもの。

  py -3 tools/checklist.py            # docs/点検メモ.md（と図）
  py -3 tools/checklist.py --xlsx     # docs/点検メモ.xlsx も出す（GUI で書く用）
  py -3 tools/checklist.py --print    # 印刷用の A4（docs/点検表/*.svg）も出す

**書き込んだものは作り直しても消えない。**「済」と「気づいたこと」を番号で引き継ぐ。
記号を足したら行が増えるだけ。

**どこから読むか。** `.xlsx` があって `.md` より新しければ Excel から、
そうでなければ `.md` から読む。**どちらから読んだかを必ず表示する。**

> **`.xlsx` は git に入れない**（`.gitignore`）。バイナリなので差分が出ず、
> 2台のPCで別々に書き込むと突き合わせができない。**git に載るのは `.md` のほう。**
> 相手のPCでは `--xlsx` を叩き直せば同じものができる。

> **表の中で `|` を素で書かない**（`.md` のとき）。列の区切りと見分けが付かず行が壊れる。
> 書きたいときは `\\|` とする。

直したら該当行を空に戻す。控えを残したいものは
[docs/測定メモ.md](../docs/測定メモ.md) へ。
"""
import argparse
import os
import re
import sys
from xml.sax.saxutils import escape

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import paths as P                                           # noqa: E402
import svg_elmt as S                                        # noqa: E402

sys.stdout.reconfigure(encoding="utf-8")

DOCS = os.path.join(P.REPO, "docs")
SYM = os.path.join(DOCS, "images", "sym")            # 1個ずつの図（.md 用）
MEMO = os.path.join(DOCS, "点検メモ.md")
XLSX = os.path.join(DOCS, "点検メモ.xlsx")
PNG = os.path.join(P.RENDER, "png")                  # Excel に貼る図（git 管理外）

# Excel の列。**番号の列を動かさない。**書き込みを引き継ぐときの手がかり。
# 節は列に持たず、**ブラウザのページと同じように見出しの行**にする
COLS = [("図", 17), ("番号", 14), ("名称", 40), ("済", 7), ("気づいたこと", 68)]
C_NUM, C_DONE, C_MEMO = 2, 4, 5                      # 1始まり
FONT = "Meiryo"          # 和名を出すので日本語を持つもの。Arial だと豆腐になる
MONO = "Consolas"        # 番号は等幅にすると桁が揃って引きやすい

# 見た目はブラウザのページに合わせる（docs/ツール.md）
SZ_BODY, SZ_HEAD, SZ_SEC = 12, 12, 14
CLR_RULE = "E6E6E6"      # 行の下の罫線
CLR_SEC = "333333"       # 節の見出しの下線
CLR_EDIT = "FFFDF2"      # 書き込む欄（ページの textarea と同じ）
CLR_HAS = "FFF8E0"       # 書いてある行（ページの tr.has と同じ）

PX_PER_UNIT = .8                # 表の中での図の大きさ
PX_MIN, PX_MAX = 22, 74         # 行が潰れる／伸びすぎるのを防ぐ

SHEET = "点検メモ"

HEAD = """# 点検メモ

記号を使っていて気づいたことを書き留める表。**`elements/` のぜんぶ**が載っている。
番号を引くだけなら [チートシート](チートシート.svg)、
姿と諸元を見るなら [図記号の一覧](sym/index.html)。

> **この表は作り直しても消えない。** `py -3 tools/checklist.py` は
> 既にある「済」と「気づいたこと」を番号で引き継ぐ。記号を足せば行が増えるだけ。

## Excel で書く

記号の姿を見ながら書ける。**書き込んだあと引数なしで叩き直すとこの表に写る。**

```bash
py -3 tools/checklist.py --xlsx    # docs/点検メモ.xlsx を作る
py -3 tools/checklist.py           # Excel の書き込みをこの表に写す
```

**`.xlsx` は git に入れていない。** バイナリなので差分が出ず、2台のPCで
別々に書き込むと突き合わせができないため。**git に載るのはこの `.md` のほう。**

## もう1台のPCで書くとき

**Excel ファイルをコピーしない。** 持ち歩くのは `.md` で、Excel は現地で組み直す。

```bash
git pull
py -3 tools/checklist.py --xlsx    # .md の書き込みごと Excel が組み上がる
```

書いたら写して送る。

```bash
py -3 tools/checklist.py
git add -A && git commit -m "点検メモ" && git push
```

> **Excel に書いたら、`git pull` の前に写す。** 写さずに pull すると `.md` の
> ほうが新しくなる。**そうなったら止まる**ので黙って消えることはないが、
> どちらを採るか選ぶ手間が出る（`--from xlsx` / `--from md`）。

`.xlsx` と `.md` の**新しいほうから読む**（どちらから読んだかは実行時に出る）。
**両方に食い違う書き込みがあるときは選ばずに止まる。**

## 書き方

- 「済」に `✔` を入れると見たしるしになる。直したら行を空に戻す
- この `.md` を直接書くときは **`|` を素で書かない。** 列の区切りと見分けが
  付かず行が壊れる。`\\|` とする（Excel 側は素の `|` でよい。写すときに逃がす）
- 残しておきたい控えは [測定メモ.md](測定メモ.md) へ

印刷して紙に書きたいときは `py -3 tools/checklist.py --print` で
`docs/点検表/` に A4 が出る。

"""

# --- Excel（--xlsx のときだけ） ---------------------------------------------
#
# **git に入れない。**バイナリで差分が出ず、2台のPCで別々に書き込むと
# 突き合わせができない。載るのは `.md` のほうで、Excel は書くための面。

XL_K = .30               # PNG を Excel に貼るときの縮尺
XL_HMIN, XL_HMAX = 26, 76        # 貼ったあとの高さ（px）の下限・上限
XL_WMAX = 230                    # 幅の上限（px）。横長の記号で列が広がるのを防ぐ
XL_ROWMIN = 34                   # 行の高さの下限（pt）。文字が12ptなので窮屈にしない

USAGE = [
    ("この表の使い方", True),
    ("", False),
    ("記号を使っていて気づいたことを、黄色の2列に書きます。", False),
    ("", False),
    ("済　　　　　見たしるし。プルダウンから ✔ を選ぶ", False),
    ("気づいたこと　直したいところ・迷ったところを書く", False),
    ("", False),
    ("書いたら閉じて、下を叩くと docs/点検メモ.md に写ります。", False),
    ("git に載るのは .md のほうです（.xlsx はバイナリなので差分が出ません）。", False),
    ("", False),
    ("    py -3 tools/checklist.py", False),
    ("", False),
    ("叩き直しても書き込みは消えません。番号で引き継ぎます。", False),
    ("記号を足したときは行が増えるだけです。", False),
    ("", False),
    ("書き方の例", True),
    ("済 = ✔ ／ 気づいたこと = 「可動接片が規格票より寝ている。07-25-01 と揃える」", False),
    ("直したら、その行を空に戻します。", False),
]


def build_xlsx(items, keep):
    """**ブラウザのページと同じ体裁にする。**

    節は列ではなく**見出しの行**にし、文字は 12pt、罫線は淡く、
    書き込む欄だけ色を敷く。表としての並べ替えより**読みやすさ**を採る
    （176行を上から順に見ていく使い方なので、絞り込みより見出しが効く）。
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import render_elmt as R

    os.makedirs(PNG, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "使い方"
    ws.column_dimensions["A"].width = 96
    for i, (line, bold) in enumerate(USAGE, 1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT, size=SZ_SEC if bold else SZ_BODY, bold=bold)
        ws.row_dimensions[i].height = 22 if bold else 18

    ws = wb.create_sheet(SHEET)
    head = Font(name=FONT, size=SZ_HEAD, bold=True)
    body = Font(name=FONT, size=SZ_BODY)
    num_f = Font(name=MONO, size=SZ_BODY)
    sec_f = Font(name=FONT, size=SZ_SEC, bold=True)
    # **書き込む欄に色を敷く。**どこを触るのか分からない表は使われない
    edit = PatternFill("solid", fgColor=CLR_EDIT)
    has = PatternFill("solid", fgColor=CLR_HAS)
    rule = Border(bottom=Side(style="thin", color=CLR_RULE))
    under = Border(bottom=Side(style="medium", color=CLR_SEC))
    wrap = Alignment(vertical="center", wrap_text=True)
    mid = Alignment(vertical="center")
    ctr = Alignment(horizontal="center", vertical="center")
    ncols = len(COLS)
    last = get_column_letter(ncols)

    for i, (name, width) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font, c.alignment, c.border = head, ctr, under
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    dv = DataValidation(type="list", formula1='"✔"', allow_blank=True)
    ws.add_data_validation(dv)

    r, wmax = 2, 0
    for ja, path in items:
        if ja:
            # 節の見出し。**ブラウザのページの h2 と同じ**（太字＋下線）
            ws.merge_cells("A%d:%s%d" % (r, last, r))
            c = ws.cell(row=r, column=1, value=ja)
            c.font, c.alignment = sec_f, Alignment(vertical="bottom")
            for i in range(1, ncols + 1):
                ws.cell(row=r, column=i).border = under
            ws.row_dimensions[r].height = 26
            r += 1

        base = os.path.basename(path)[:-5]
        _, _, name, _ = S.body(path, False)
        done, memo = keep.get(base, ("", ""))
        written = bool(done or memo.strip())
        for col, val, al, ft in ((2, base, mid, num_f), (3, name, wrap, body),
                                 (4, done, ctr, body), (5, memo, wrap, body)):
            c = ws.cell(row=r, column=col, value=val)
            c.font, c.alignment, c.border = ft, al, rule
            if col in (C_DONE, C_MEMO):
                c.fill = edit
            if written:
                c.fill = has
        ws.cell(row=r, column=1).border = rule
        dv.add(ws.cell(row=r, column=C_DONE))

        png = os.path.join(PNG, base + ".png")
        R.draw_one(path, marks=False, caption=False, pad=8).save(png)
        img = XLImage(png)
        k = XL_K
        if img.height * k > XL_HMAX:
            k = XL_HMAX / img.height
        if img.height * k < XL_HMIN:
            k = XL_HMIN / img.height
        if img.width * k > XL_WMAX:                  # 横長は幅で頭打ちにする
            k = XL_WMAX / img.width
        img.width, img.height = int(img.width * k), int(img.height * k)
        wmax = max(wmax, img.width)
        ws.add_image(img, "A%d" % r)
        ws.row_dimensions[r].height = max(XL_ROWMIN, img.height * .75 + 8)
        r += 1

    ws.column_dimensions["A"].width = wmax / 7. + 2.5
    ws.sheet_view.showGridLines = False       # 罫線は自分で引くので方眼は消す
    wb.save(XLSX)
    return len(items)


# --- 印刷用 A4（--print のときだけ） ---------------------------------------
PW, PH, MARGIN = 210., 297., 9.
HEAD_H, FOOT_H = 11., 6.
X_CHK, X_FIG, X_NAME, X_MEMO = 14., 44., 94., PW - MARGIN
MM_PER_UNIT = .20
MAX_K = .38
ROW_MIN, ROW_PAD = 11., 3.
TITLE = "図記号 点検メモ"
NOTE = "気づいたことをその場で書く。直したら docs/点検メモ.md に写す"


def rows():
    """[(節の表示名 or None, パス)] を elements/ の並びのまま返す

    節が変わる行にだけ表示名を入れる（見出しを出す合図）。
    """
    secs = S.sections()
    out, cur = [], None
    for key in sorted(secs):
        ja = P.dirname_ja(os.path.join(P.ELEMENTS, *key.split("/")))
        for p in sorted(secs[key]):
            out.append((ja if ja != cur else None, p))
            cur = ja
    return out


def md_escape(s):
    """表のセルに入れられる形にする

    **Excel には縦棒も改行も素で書ける。** それをそのまま `.md` に流すと
    列が1つ増えて行が壊れ、次に読むときに名称とメモがずれる（実際にやった）。
    書き出す側で必ず逃がす。
    """
    return (s.replace("\\", "\\\\").replace("|", "\\|")
            .replace("\r\n", "<br>").replace("\n", "<br>").replace("\r", "<br>"))


def md_unescape(s):
    return (s.replace("<br>", "\n").replace("\\|", "|").replace("\\\\", "\\"))


def read_memo():
    """既にある点検メモから {番号: (済, 気づいたこと)} を読む

    **手で書き換えられている前提で読む。** 列の幅も空白も揃っていないので、
    番号のセルだけを頼りに引く。読めない行は捨てずに黙って飛ばす
    （消えたら書いた人が困るのは同じだが、壊れた行に引きずられて
    全部を落とすほうが害が大きい）。
    """
    if not os.path.isfile(MEMO):
        return {}
    out = {}
    for line in open(MEMO, encoding="utf-8"):
        if not line.lstrip().startswith("|"):
            continue
        # **`\|` で割らない。** 表の中に縦棒を書けるようにエスケープを勧めておいて
        # ここで素の区切りとして割ると、書いた本人のメモが途中で切れる（実際にやった）
        cells = [c.strip() for c in
                 re.split(r"(?<!\\)\|", line.strip().strip("|"))]
        if len(cells) < 5:
            continue
        m = re.fullmatch(r"`([0-9A-Za-z_\-]+)`", cells[1])
        if m:
            out[m.group(1)] = (md_unescape(cells[3]), md_unescape(cells[4]))
    return out


def read_xlsx():
    """Excel から {番号: (済, 気づいたこと)} を読む"""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, data_only=True)
    ws = wb[SHEET] if SHEET in wb.sheetnames else wb.worksheets[0]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < C_MEMO or not row[C_NUM - 1]:
            continue
        num = str(row[C_NUM - 1]).strip()
        done = str(row[C_DONE - 1] or "").strip()
        memo = str(row[C_MEMO - 1] or "").strip()
        out[num] = (done, memo)
    return out


def has_note(v):
    return bool(v[0].strip() or v[1].strip())


def conflicts(a, b):
    """2つの書き込みのうち、**片方にしかない／食い違う**番号を返す"""
    out = []
    for k in sorted(set(a) | set(b)):
        x, y = a.get(k, ("", "")), b.get(k, ("", ""))
        if (has_note(x) or has_note(y)) and x != y:
            out.append(k)
    return out


def read_notes(force=None):
    """書き込みを読む。**どちらから読んだかも返す**

    `.xlsx` があって `.md` より新しければ Excel が勝つ。
    片方だけを直したときに、古いほうで上書きしてしまわないため。

    **ただし新しさだけでは足りない。** 家で Excel に書いたまま写さずに
    `git pull` すると `.md` のほうが新しくなり、**Excel の書き込みが
    黙って消える。** 両方に食い違う書き込みがあるときは選ばずに返し、
    呼び手に止めさせる。
    """
    has_x, has_m = os.path.isfile(XLSX), os.path.isfile(MEMO)
    x = read_xlsx() if has_x else {}
    m = read_memo() if has_m else {}
    if force == "xlsx" and has_x:
        return x, "点検メモ.xlsx", []
    if force == "md" and has_m:
        return m, "点検メモ.md", []
    if not has_x:
        return m, "点検メモ.md" if has_m else "（まだ無い）", []
    if not has_m:
        return x, "点検メモ.xlsx", []
    bad = conflicts(x, m)
    if os.path.getmtime(XLSX) > os.path.getmtime(MEMO):
        return x, "点検メモ.xlsx", bad
    return m, "点検メモ.md", bad


def sym_svg(path):
    """1個ずつの図を書き出して (ファイル名, 表示する高さpx) を返す"""
    base = os.path.basename(path)[:-5]
    _, (x0, y0, x1, y1), _, _ = S.body(path, False)
    open(os.path.join(SYM, base + ".svg"), "w", encoding="utf-8").write(
        S.one(path))
    # **高さ0の記号がある**（07-01-03 は横棒1本）。そのまま使うと表示が潰れる
    h = (y1 - y0) or (x1 - x0)
    return base, max(PX_MIN, min(PX_MAX, int(round(h * PX_PER_UNIT))))


def build_md(items, keep):
    out = [HEAD]
    cur = None
    for ja, path in items:
        if ja and ja != cur:
            cur = ja
            out.append("\n## %s\n" % ja)
            out.append("| 図 | 番号 | 名称 | 済 | 気づいたこと |")
            out.append("|---|---|---|---|---|")
        base, px = sym_svg(path)
        _, _, name, _ = S.body(path, False)
        done, memo = keep.get(base, ("", ""))
        out.append('| <img src="sym/svg/%s.svg" height="%d"> | `%s` | %s | %s | %s |'
                   % (base, px, base, md_escape(name),
                      md_escape(done), md_escape(memo)))
    return "\n".join(out) + "\n"


# --- 以下 --print の印刷用 --------------------------------------------------

def row_svg(path, top, h):
    """1行ぶん。**図はセルいっぱいに引き伸ばす**（上限 MAX_K）

    行の高さは記号の実寸で決めるが、図はセルに合わせて拡大する。
    そうしないと 07-01-03（横棒1本）が 4mm に潰れて紙の上で何か分からなくなる。
    横棒・縦棒だけの記号は広がりが片方 0 になるので、割る前に避ける。
    """
    g, (x0, y0, x1, y1), ja, _ = S.body(path, False)
    figw, figh = X_FIG - X_CHK - 4, h - 2
    k = MAX_K
    if x1 - x0 > 0:
        k = min(k, figw / (x1 - x0))
    if y1 - y0 > 0:
        k = min(k, figh / (y1 - y0))
    cx, cy = (x0 + x1) / 2, (y0 + y1) / 2
    base = os.path.basename(path)[:-5]
    return [
        '<g transform="translate(%g %g) scale(%g) translate(%g %g)"'
        ' stroke="black" fill="none">%s</g>'
        % ((X_CHK + X_FIG) / 2, top + h / 2, k, -cx, -cy, g),
        '<rect x="%g" y="%g" width="3.4" height="3.4" fill="none"'
        ' stroke="#666" stroke-width="0.25"/>' % (MARGIN + 0.6, top + h / 2 - 1.7),
        '<text x="%g" y="%g" font-family="%s" font-size="2.6" fill="black">%s</text>'
        % (X_FIG + 2, top + h / 2 - 0.3, S.CAPFONT, escape(base)),
        '<text x="%g" y="%g" font-family="%s" font-size="2.2" fill="#333">%s</text>'
        % (X_FIG + 2, top + h / 2 + 3.2, S.CAPFONT,
           escape(S._fit(ja, X_NAME - X_FIG - 3, 2.2))),
        '<line x1="%g" y1="%g" x2="%g" y2="%g" stroke="#ccc" stroke-width="0.15"/>'
        % (X_NAME, top + 0.5, X_NAME, top + h - 0.5),
        '<line x1="%g" y1="%g" x2="%g" y2="%g" stroke="#999" stroke-width="0.2"/>'
        % (MARGIN, top + h, X_MEMO, top + h),
    ]


def band(ja, top):
    return [
        '<rect x="%g" y="%g" width="%g" height="5" fill="#eee"/>'
        % (MARGIN, top, X_MEMO - MARGIN),
        '<text x="%g" y="%g" font-family="%s" font-size="2.9" font-weight="bold"'
        ' fill="black">%s</text>' % (MARGIN + 1.5, top + 3.6, S.CAPFONT, escape(ja)),
    ]


def page(parts, n, total):
    head = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<svg xmlns="http://www.w3.org/2000/svg" width="%gmm" height="%gmm"'
        ' viewBox="0 0 %g %g">' % (PW, PH, PW, PH),
        '<rect width="%g" height="%g" fill="white"/>' % (PW, PH),
        '<g stroke-linecap="round" stroke-linejoin="round">',
        '<text x="%g" y="%g" font-family="%s" font-size="4.4" font-weight="bold"'
        ' fill="black">%s</text>' % (MARGIN, MARGIN + 3.4, S.CAPFONT, escape(TITLE)),
        # **節名はここに並べない。**5節に跨るページがあり見出しに食い込む。
        # どの節かは行のあいだの帯が持つ
        '<text x="%g" y="%g" text-anchor="end" font-family="%s" font-size="2.8"'
        ' fill="#555">%s</text>'
        % (X_MEMO, MARGIN + 3.4, S.CAPFONT, escape("%d / %d" % (n, total))),
        '<text x="%g" y="%g" font-family="%s" font-size="2.3" fill="#666">%s</text>'
        % (MARGIN, MARGIN + 7.4, S.CAPFONT, escape(NOTE)),
        '<line x1="%g" y1="%g" x2="%g" y2="%g" stroke="#333" stroke-width="0.4"/>'
        % (MARGIN, MARGIN + HEAD_H - 1.4, X_MEMO, MARGIN + HEAD_H - 1.4),
        '<text x="%g" y="%g" font-family="%s" font-size="2.2" fill="#666">番号・名称</text>'
        % (X_FIG + 2, MARGIN + HEAD_H + 1.4, S.CAPFONT),
        '<text x="%g" y="%g" font-family="%s" font-size="2.2" fill="#666">気づいたこと</text>'
        % (X_NAME + 2, MARGIN + HEAD_H + 1.4, S.CAPFONT),
    ]
    return "\n".join(head + parts + ['</g></svg>', ''])


def build_print(items):
    d = os.path.join(DOCS, "点検表")
    os.makedirs(d, exist_ok=True)
    for old in os.listdir(d):                    # 数が減ったとき古いページを残さない
        if old.endswith(".svg"):
            os.remove(os.path.join(d, old))
    bottom = PH - MARGIN - FOOT_H
    pages, parts = [], []
    y = MARGIN + HEAD_H + 2.4
    cur = ""
    for ja, path in items:
        if ja:
            cur = ja
        _, (_, y0, _, y1), _, _ = S.body(path, False)
        h = max(ROW_MIN, (y1 - y0) * MM_PER_UNIT + ROW_PAD)
        show, cont = ja is not None, False
        if y + h + (5.4 if show else 0) > bottom and parts:
            pages.append(parts)
            parts, y = [], MARGIN + HEAD_H + 2.4
            # **送った先で節名を出し直す。** でないとそのページだけ見たとき
            # どの節を見ているのか分からない
            show, cont = True, ja is None
        if show:
            parts += band(cur + ("（つづき）" if cont else ""), y)
            y += 5.4
        parts += row_svg(path, y, h)
        y += h
    if parts:
        pages.append(parts)
    for i, parts in enumerate(pages, 1):
        open(os.path.join(d, "%02d.svg" % i), "w", encoding="utf-8").write(
            page(parts, i, len(pages)))
    return len(pages)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", action="store_true",
                    help="Excel も出す（GUI で書く用）")
    ap.add_argument("--print", dest="pr", action="store_true",
                    help="印刷用の A4 も出す")
    ap.add_argument("--from", dest="src", choices=("md", "xlsx"),
                    help="食い違ったとき、どちらを採るか決める")
    a = ap.parse_args()

    os.makedirs(SYM, exist_ok=True)
    items = rows()
    keep, src, bad = read_notes(a.src)

    if bad:
        # **黙ってどちらかで上書きしない。** 家で Excel に書いたまま写さずに
        # git pull すると .md が新しくなり、書き込みが消える。実際に起きうる
        print("**食い違っています。**どちらを採るか決めてください。")
        print("  .md と .xlsx で中身の違う行: %s" % "・".join(bad[:12])
              + ("　ほか%d件" % (len(bad) - 12) if len(bad) > 12 else ""))
        print("  Excel に書いたぶんを採る : py -3 tools/checklist.py --from xlsx")
        print("  .md のぶんを採る         : py -3 tools/checklist.py --from md")
        return 1

    written = sum(1 for v in keep.values() if has_note(v))
    print("読み取り: %s（書き込み %d行）" % (src, written))

    open(MEMO, "w", encoding="utf-8").write(build_md(items, keep))
    print("書き出し: %s  %d行" % (os.path.relpath(MEMO, P.REPO), len(items)))

    # **既にあるなら黙って作り直す。**片方だけ新しい状態を残すと、
    # 次に叩いたとき古いほうから読んで書き込みが消える
    if a.xlsx or os.path.isfile(XLSX):
        try:
            build_xlsx(items, keep)
            print("書き出し: %s" % os.path.relpath(XLSX, P.REPO))
        except PermissionError:
            # **Excel で開いたままだと書けない。**生の例外を出さない。
            # `.md` は先に書けているので、読み取った中身は失われていない
            print("**%s を Excel で開いたままです。**閉じてから叩き直してください。"
                  % os.path.relpath(XLSX, P.REPO))
            print("（%s は書けているので、書き込みは失われていません）"
                  % os.path.relpath(MEMO, P.REPO))
            return 1

    if a.pr:
        n = build_print(items)
        print("印刷用: docs/点検表/  A4 %d枚" % n)
    return 0


if __name__ == "__main__":
    sys.exit(main())
